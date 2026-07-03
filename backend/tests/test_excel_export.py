from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from app.db.models import (
    AcademicPeriod,
    Base,
    Competency,
    Course,
    CourseCompetency,
    Curriculum,
    EvaluationCriterion,
    EvaluationResult,
    Program,
)
from app.services.excel_export import export_periods_to_excel


def _session() -> Session:
    engine = create_engine("sqlite:///:memory:", future=True)
    Base.metadata.create_all(bind=engine)
    return sessionmaker(bind=engine, future=True)()


def test_excel_export_uses_separate_manual_sheets():
    """El reporte debe mantener las hojas prometidas por el manual de usuario."""
    db = _session()

    # Arrange
    program = Program(name="Ingenieria Civil Computacion")
    db.add(program)
    db.flush()
    curriculum = Curriculum(
        program_id=program.id,
        year=2027,
        version="PE 2027 COMPUTACION",
        display_name="PE 2027 Computacion",
    )
    db.add(curriculum)
    db.flush()
    course = Course(
        curriculum_id=curriculum.id,
        code="ICC101",
        title="Introduccion a la Ingenieria",
        sort_order=1,
    )
    competency = Competency(
        curriculum_id=curriculum.id,
        code="TIC1",
        group="TIC",
        description="Disena soluciones de software.",
        sort_order=1,
    )
    db.add_all([course, competency])
    db.flush()
    criterion = EvaluationCriterion(
        competency_id=competency.id,
        name="Evidencia para TIC1",
        description=competency.description,
        threshold=0.22,
    )
    db.add(criterion)
    db.flush()
    db.add(CourseCompetency(course_id=course.id, competency_id=competency.id))
    period = AcademicPeriod(name="2027-1", curriculum_id=curriculum.id, status="ready")
    db.add(period)
    db.flush()
    db.add(
        EvaluationResult(
            period_id=period.id,
            document_id=None,
            criterion_id=criterion.id,
            course_id=course.id,
            score=82,
            confidence=0.82,
            status="ready",
        )
    )
    db.commit()

    # Act
    workbook_bytes = export_periods_to_excel(db, [period.id])
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True)

    # Assert
    assert workbook.sheetnames == [
        "Resumen",
        "Mapa de Calor",
        "Detalle por Curso",
        "Detalle por Competencia",
        "Brechas",
    ]
    assert workbook["Resumen"]["A1"].value.startswith("Reporte de Validacion")
    assert workbook["Mapa de Calor"]["A1"].value == "Ramo / Competencia"
    assert workbook["Detalle por Curso"]["A1"].value == "Curso"
